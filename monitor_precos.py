"""
monitor_precos.py
1. Baixa a planilha do Google Drive
2. Raspa preços do ML via API oficial e do fornecedor via scraping
3. Atualiza colunas 4-6 e colore status
4. Sobe a planilha atualizada de volta ao Google Drive
5. Envia alertas via Telegram quando necessário
"""
 
import re
import json
import time
import logging
import requests
import os
from datetime import datetime
from pathlib import Path
 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup
 
from config import (
    TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID,
    COMISSAO_ML, IMPOSTO_DAS, MARGEM_MIN, FRETE_FIXO,
    GDRIVE_FILE_ID, PLANILHA_PATH,
)
 
# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler("/data/monitor.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9",
}
 
DESCONTO = COMISSAO_ML + IMPOSTO_DAS + MARGEM_MIN  # 0.28
 
# Token ML — lido da variável de ambiente ou config
ML_ACCESS_TOKEN = os.environ.get("ML_ACCESS_TOKEN", "")
 
# ── Fórmula PMC ──────────────────────────────────────────────────────────────
def calcular_pmc(preco_ml: float) -> float:
    return round(preco_ml * (1 - DESCONTO) - FRETE_FIXO, 2)
 
# ── Google Drive ──────────────────────────────────────────────────────────────
def gdrive_download(file_id: str, destino: str) -> bool:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    try:
        sess = requests.Session()
        resp = sess.get(url, stream=True, timeout=30)
        for key, val in resp.cookies.items():
            if "download_warning" in key:
                url = f"{url}&confirm={val}"
                resp = sess.get(url, stream=True, timeout=30)
                break
        resp.raise_for_status()
        Path(destino).parent.mkdir(parents=True, exist_ok=True)
        with open(destino, "wb") as f:
            for chunk in resp.iter_content(32768):
                f.write(chunk)
        log.info("Planilha baixada do Drive → %s", destino)
        return True
    except Exception as e:
        log.error("Erro ao baixar do Drive: %s", e)
        return False
 
 
def gdrive_upload(file_id: str, caminho: str) -> bool:
    log.info("Planilha atualizada salva em: %s", caminho)
    log.info("Para sincronizar com o Drive, use: rclone copy %s gdrive:/Monitor-Precos/", caminho)
    return True
 
 
# ── Scraper ML via API Oficial ────────────────────────────────────────────────
def extrair_item_id_ml(url: str) -> tuple[str, str]:
    """
    Retorna (tipo, id) onde tipo é 'item' (MLB123) ou 'catalog' (MLB123 de catálogo).
    """
    # Produto de catálogo: /p/MLB123
    m = re.search(r'/p/(MLB\d+)', url, re.IGNORECASE)
    if m:
        return ('catalog', m.group(1).upper())
 
    # Item direto: /MLB123... ou MLB-123
    m = re.search(r'(MLB-?\d+)', url, re.IGNORECASE)
    if m:
        return ('item', m.group(1).upper().replace("-", ""))
 
    return (None, None)
 
 
def extrair_preco_ml(url: str) -> float | None:
    try:
        tipo, item_id = extrair_item_id_ml(url)
        if not item_id:
            log.warning("Item ID ML não encontrado na URL: %s", url)
            return None
 
        headers_ml = {"Authorization": f"Bearer {ML_ACCESS_TOKEN}"} if ML_ACCESS_TOKEN else {}
 
        if tipo == 'catalog':
            # Busca itens do catálogo e pega o menor preço original
            api_url = f"https://api.mercadolibre.com/products/{item_id}/items"
            resp = requests.get(api_url, headers=headers_ml, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                resultados = data.get("results", [])
                precos = []
                for r in resultados:
                    p = r.get("original_price") or r.get("price")
                    if p:
                        precos.append(float(p))
                if precos:
                    preco = min(precos)
                    log.info("API ML catálogo: %s → R$ %.2f", item_id, preco)
                    return preco
 
            # Fallback: busca o item diretamente
            api_url = f"https://api.mercadolibre.com/items/{item_id}"
            resp = requests.get(api_url, headers=headers_ml, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                preco = data.get("original_price") or data.get("price")
                if preco:
                    log.info("API ML item direto: %s → R$ %.2f", item_id, float(preco))
                    return float(preco)
        else:
            api_url = f"https://api.mercadolibre.com/items/{item_id}"
            resp = requests.get(api_url, headers=headers_ml, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                preco = data.get("original_price") or data.get("price")
                if preco:
                    log.info("API ML: %s → R$ %.2f", item_id, float(preco))
                    return float(preco)
 
        log.warning("Preço ML não encontrado: %s (status %s)", item_id, resp.status_code)
        return None
 
    except Exception as e:
        log.error("Erro ML (%s): %s", url, e)
        return None
 
 
# ── Scraper Fornecedor ────────────────────────────────────────────────────────
def extrair_preco_fornecedor(url: str) -> float | None:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
 
        # Meta tags og
        for meta_name in ["product:price:amount", "og:price:amount"]:
            tag = soup.find("meta", property=meta_name)
            if tag and tag.get("content"):
                return float(re.sub(r"[^\d.]", "", tag["content"].replace(",", ".")))
 
        # itemprop=price
        tag = soup.find(attrs={"itemprop": "price"})
        if tag:
            valor = tag.get("content") or tag.get_text(strip=True)
            valor = re.sub(r"[^\d.,]", "", valor).replace(".", "").replace(",", ".")
            if valor:
                return float(valor)
 
        # JSON-LD
        for tag in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(tag.string)
                items = data if isinstance(data, list) else [data]
                for item in items:
                    offers = item.get("offers", {})
                    if isinstance(offers, list):
                        offers = offers[0]
                    price = offers.get("price")
                    if price:
                        return float(price)
            except Exception:
                pass
 
        # Seletores comuns de preço
        seletores = [
            "[class*='price'] [class*='value']",
            "[class*='preco']",
            "[class*='price']",
            "[id*='price']",
        ]
        for sel in seletores:
            el = soup.select_one(sel)
            if el:
                texto = el.get_text(strip=True)
                nums = re.sub(r"[^\d,]", "", texto)
                if nums:
                    valor = nums.replace(",", ".")
                    try:
                        v = float(valor)
                        if v > 0:
                            return v
                    except Exception:
                        pass
 
        log.warning("Preço fornecedor não encontrado: %s", url)
        return None
    except Exception as e:
        log.error("Erro fornecedor (%s): %s", url, e)
        return None
 
 
# ── Telegram ─────────────────────────────────────────────────────────────────
def telegram_send(mensagem: str) -> None:
    if not TELEGRAM_BOT_TOKEN or TELEGRAM_BOT_TOKEN == "SEU_TOKEN_AQUI":
        log.warning("Telegram não configurado.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": mensagem, "parse_mode": "HTML"}
    try:
        r = requests.post(url, json=payload, timeout=10)
        r.raise_for_status()
        log.info("Telegram: mensagem enviada.")
    except Exception as e:
        log.error("Erro Telegram: %s", e)
 
 
# ── Planilha ──────────────────────────────────────────────────────────────────
HEADER_ROW = 1
DATA_START  = 2
 
COL_SKU, COL_LINK_ML, COL_LINK_FORN = 1, 2, 3
COL_PRECO_ML, COL_PRECO_FORN, COL_PMC = 4, 5, 6
COL_STATUS, COL_ATUALIZADO = 7, 8
 
COR_OK     = "C6EFCE"
COR_ALERTA = "FFEB9C"
COR_PERIGO = "FFC7CE"
COR_HEADER = "2E75B6"
 
 
def garantir_cabecalhos(ws) -> None:
    if ws.cell(1, 1).value == "SKU":
        return
    headers = [
        "SKU", "Link ML", "Link Fornecedor",
        "Preço ML (R$)", "Preço Fornecedor (R$)",
        "PMC Máximo (R$)", "Status", "Última Atualização",
    ]
    for col, titulo in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center")
    larguras = {"A":18,"B":55,"C":55,"D":18,"E":22,"F":20,"G":30,"H":22}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
 
 
def colorir_linha(ws, row: int, cor: str) -> None:
    fill = PatternFill("solid", fgColor=cor)
    for col in range(COL_PRECO_ML, COL_ATUALIZADO + 1):
        ws.cell(row=row, column=col).fill = fill
 
 
def estados_anteriores(ws) -> dict:
    estados = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        sku    = row[COL_SKU - 1]
        status = row[COL_STATUS - 1] if len(row) >= COL_STATUS else None
        if sku:
            estados[str(sku)] = status or ""
    return estados
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
def processar() -> None:
    if not gdrive_download(GDRIVE_FILE_ID, PLANILHA_PATH):
        log.error("Não foi possível baixar a planilha. Abortando.")
        return
 
    wb = openpyxl.load_workbook(PLANILHA_PATH)
    ws = wb.active
    garantir_cabecalhos(ws)
 
    ant = estados_anteriores(ws)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    alertas = []
 
    for row_idx in range(DATA_START, ws.max_row + 1):
        sku       = ws.cell(row_idx, COL_SKU).value
        link_ml   = ws.cell(row_idx, COL_LINK_ML).value
        link_forn = ws.cell(row_idx, COL_LINK_FORN).value
 
        if not sku or not link_ml or not link_forn:
            continue
 
        log.info("Processando SKU %s ...", sku)
 
        preco_ml   = extrair_preco_ml(str(link_ml))
        time.sleep(1.5)
        preco_forn = extrair_preco_fornecedor(str(link_forn))
        time.sleep(1.5)
 
        ws.cell(row_idx, COL_ATUALIZADO).value = agora
 
        if preco_ml is None or preco_forn is None:
            ws.cell(row_idx, COL_STATUS).value = "⚠️ Erro na leitura"
            colorir_linha(ws, row_idx, COR_ALERTA)
            continue
 
        pmc = calcular_pmc(preco_ml)
 
        ws.cell(row_idx, COL_PRECO_ML).value   = preco_ml
        ws.cell(row_idx, COL_PRECO_FORN).value = preco_forn
        ws.cell(row_idx, COL_PMC).value        = pmc
 
        for col in (COL_PRECO_ML, COL_PRECO_FORN, COL_PMC):
            ws.cell(row_idx, col).number_format = 'R$ #,##0.00'
 
        status_ant = ant.get(str(sku), "")
 
        if preco_forn > pmc:
            status = "🚨 ACIMA DO PMC"
            colorir_linha(ws, row_idx, COR_PERIGO)
            if status_ant != status:
                alertas.append(
                    f"🚨 <b>ALERTA — Fornecedor acima do PMC</b>\n"
                    f"SKU: <code>{sku}</code>\n"
                    f"Preço ML: R$ {preco_ml:,.2f}\n"
                    f"PMC Máximo: R$ {pmc:,.2f}\n"
                    f"Preço Fornecedor: R$ {preco_forn:,.2f}  ❌\n"
                    f"Data: {agora}"
                )
        else:
            status = "✅ OK"
            colorir_linha(ws, row_idx, COR_OK)
            if status_ant == "🚨 ACIMA DO PMC":
                alertas.append(
                    f"✅ <b>NORMALIZADO — Fornecedor voltou abaixo do PMC</b>\n"
                    f"SKU: <code>{sku}</code>\n"
                    f"Preço ML: R$ {preco_ml:,.2f}\n"
                    f"PMC Máximo: R$ {pmc:,.2f}\n"
                    f"Preço Fornecedor: R$ {preco_forn:,.2f}  ✅\n"
                    f"Data: {agora}"
                )
 
        ws.cell(row_idx, COL_STATUS).value = status
        log.info("  ML=R$%.2f  Forn=R$%.2f  PMC=R$%.2f  → %s",
                 preco_ml, preco_forn, pmc, status)
 
    wb.save(PLANILHA_PATH)
    log.info("Planilha salva localmente.")
    gdrive_upload(GDRIVE_FILE_ID, PLANILHA_PATH)
 
    for alerta in alertas:
        telegram_send(alerta)
 
    if not alertas:
        log.info("Nenhuma mudança de status detectada.")
 
 
if __name__ == "__main__":
    log.info("=== Container iniciado — aguardando agendamento do Dokploy ===")
    while True:
        time.sleep(60)
